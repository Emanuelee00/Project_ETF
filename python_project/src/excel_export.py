"""Excel export utilities (stable professional version)."""

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import ScatterChart, Reference, Series


country_colors = {
    "UK": "D9EAF7",
    "France": "FADADD",
    "Germany": "FFF4CC",
    "Italy": "DFF2DF",
    "Switzerland": "E0F7FA",
    "Netherlands": "E6E6FA",
    "Singapore": "FFE5B4",
    "Other": "FFFFFF"
}


# --------------------------------------------------
# Utilities
# --------------------------------------------------

def auto_adjust_columns(ws):
    """Auto adjust column width."""
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2


def format_metrics(ws):
    """Professional numeric formatting."""

    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2):
        for cell in row:

            if not isinstance(cell.value, (float, int)):
                continue

            header = headers[cell.column - 1]

            if header is None:
                continue

            # Percent values
            if any(k in header for k in [
                "Perf", "CAGR", "Return", "Volatility", "Drawdown"
            ]):
                cell.number_format = "0.00%"

            # Ratios
            elif any(k in header for k in [
                "Sharpe", "Beta", "Skewness", "Kurtosis", "Correlation"
            ]):
                cell.number_format = "0.00"

            # Covariance
            elif "Covariance" in header:
                cell.number_format = "0.0000"

            # Price
            elif "Price" in header:
                cell.number_format = "0.00"

def format_weights(ws):
    """Format portfolio weights as percentage."""

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (float, int)):
                cell.number_format = "0.00%"


def apply_country_colors(ws):
    """Color rows by country (if column exists)."""
    headers = [cell.value for cell in ws[1]]

    if "Country" not in headers:
        return

    country_col_index = headers.index("Country") + 1

    for row in ws.iter_rows(min_row=2):
        country = row[country_col_index - 1].value
        fill_color = country_colors.get(country, "FFFFFF")

        for cell in row:
            cell.fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid"
            )


def add_correlation_heatmap(ws):
    """Add heatmap to correlation matrix."""
    rule = ColorScaleRule(
        start_type="num", start_value=-1, start_color="F8696B",
        mid_type="num", mid_value=0, mid_color="FFEB84",
        end_type="num", end_value=1, end_color="63BE7B"
    )

    max_row = ws.max_row
    max_col = ws.max_column

    last_col_letter = get_column_letter(max_col)

    ws.conditional_formatting.add(
        f"B2:{last_col_letter}{max_row}",
        rule
    )


def add_legend(ws):
    """Add country legend dynamically on the right side."""
    start_col = ws.max_column + 2
    col_letter = get_column_letter(start_col)

    ws[f"{col_letter}2"] = "Country Legend"
    ws[f"{col_letter}2"].font = Font(bold=True)

    row = 3
    for country, color in country_colors.items():
        ws[f"{col_letter}{row}"] = country
        ws[f"{col_letter}{row}"].fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )
        row += 1

def add_seasonality_sheet(wb, weekly_seasonality):

    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Weekly Seasonality")

    if weekly_seasonality is None or len(weekly_seasonality) == 0:
        ws["A1"] = "No seasonality data available"
        return

    horizons = [2, 3, 4, 5]
    row_cursor = 1

    for horizon in horizons:

        # Titolo blocco
        ws.cell(row=row_cursor, column=1, value=f"Seasonality - Last {horizon} Years").font = Font(bold=True)
        row_cursor += 2

        # Header
        headers = ["Asset"] + [f"W{w}" for w in range(1, 53)]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_cursor, column=col, value=header)
            cell.font = Font(bold=True)

        row_cursor += 1

        for asset, pivot in weekly_seasonality.items():

            if pivot.shape[1] < horizon:
                continue

            last_years = pivot.iloc[:, -horizon:]
            avg = last_years.mean(axis=1)

            ws.cell(row=row_cursor, column=1, value=asset)

            for week in range(1, 53):

                value = avg.get(week, None)
                cell = ws.cell(row=row_cursor, column=week + 1, value=value)
                cell.number_format = "0.00%"

                if value is not None:
                    if value > 0.005:
                        fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value > 0:
                        fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                    elif value > -0.005:
                        fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
                    else:
                        fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

                    cell.fill = fill

            row_cursor += 1

        row_cursor += 3  # spazio tra blocchi

    ws.column_dimensions["A"].width = 20
    for col in range(2, 54):
        ws.column_dimensions[get_column_letter(col)].width = 8

# """ def add_efficient_frontier(ws_metrics):

#     chart = ScatterChart()
#     chart.title = "Risk vs Return"
#     chart.x_axis.title = "Volatility"
#     chart.y_axis.title = "Annual Return"

#     vol_col = None
#     ret_col = None

#     headers = [c.value for c in ws_metrics[1]]

#     vol_col = headers.index("Volatility 1Y") + 1
#     ret_col = headers.index("Annual Return") + 1

#     xvalues = Reference(ws_metrics,
#                         min_col=vol_col,
#                         min_row=2,
#                         max_row=ws_metrics.max_row)

#     yvalues = Reference(ws_metrics,
#                         min_col=ret_col,
#                         min_row=2,
#                         max_row=ws_metrics.max_row)

#     series = Series(yvalues, xvalues)
#     chart.series.append(series)

#     ws_metrics.add_chart(chart, "M2") """

# --------------------------------------------------
# Main Export Function
# --------------------------------------------------

def export_to_excel(
    metrics,
    correlation,
    max_sharpe,
    min_vol,
    weekly_seasonality,
    output_path
):
    """Export portfolio results to formatted Excel report with Dashboard."""

    wb = Workbook()

    # =====================================================
    # ================= DASHBOARD =========================
    # =====================================================

    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    ws_dash["A1"] = "Portfolio Summary"
    ws_dash["A1"].font = Font(size=14, bold=True)

    total_assets = len(metrics)
    avg_sharpe = metrics["Sharpe Ratio"].mean()
    avg_vol = metrics["Volatility 1Y"].mean()
    avg_cagr = metrics["CAGR 3Y"].mean()
    avg_dd = metrics["Max Drawdown"].mean()

    ws_dash["A3"] = "Number of Assets"
    ws_dash["B3"] = total_assets

    ws_dash["A4"] = "Average Sharpe"
    ws_dash["B4"] = avg_sharpe
    ws_dash["B4"].number_format = "0.00"

    ws_dash["A5"] = "Average Volatility"
    ws_dash["B5"] = avg_vol
    ws_dash["B5"].number_format = "0.00%"

    ws_dash["A6"] = "Average CAGR 3Y"
    ws_dash["B6"] = avg_cagr
    ws_dash["B6"].number_format = "0.00%"

    ws_dash["A7"] = "Average Max Drawdown"
    ws_dash["B7"] = avg_dd
    ws_dash["B7"].number_format = "0.00%"

    auto_adjust_columns(ws_dash)

    # =====================================================
    # ================= METRICS ===========================
    # =====================================================

    ws_metrics = wb.create_sheet("Metrics")

    for r in dataframe_to_rows(metrics, index=True, header=True):
        ws_metrics.append(r)
    
    # -----------------------------------------------------
    # ADD HYPERLINK ONLY ON ETF NAME COLUMN
    # -----------------------------------------------------

    # Header row
    headers = [cell.value for cell in ws_metrics[1]]

    # Trova colonna Name
    if "Name" in headers:
        name_col_index = headers.index("Name") + 1
    else:
        name_col_index = None

    if name_col_index:

        for row in range(2, ws_metrics.max_row + 1):

            # Colonna A = Ticker (indice)
            ticker_value = ws_metrics.cell(row=row, column=1).value

            # Colonna Name
            name_cell = ws_metrics.cell(row=row, column=name_col_index)

            if ticker_value and name_cell.value:

                url = f"http://localhost:8501/?ticker={ticker_value}"

                name_cell.hyperlink = url
                name_cell.style = "Hyperlink"

    ws_metrics.freeze_panes = "B2"

    format_metrics(ws_metrics)
    apply_country_colors(ws_metrics)
    auto_adjust_columns(ws_metrics)
    add_legend(ws_metrics)

    # Add Efficient Frontier chart to Metrics
    # if "Volatility 1Y" in metrics.columns and "Annual Return" in metrics.columns:
    #     add_efficient_frontier(ws_metrics)

    # =====================================================
    # ================= CORRELATION =======================
    # =====================================================

    ws_corr = wb.create_sheet("Correlation")

    for r in dataframe_to_rows(correlation, index=True, header=True):
        ws_corr.append(r)

    ws_corr.freeze_panes = "B2"

    auto_adjust_columns(ws_corr)
    add_correlation_heatmap(ws_corr)

    # =====================================================
    # ================= MAX SHARPE ========================
    # =====================================================

    ws_max = wb.create_sheet("Max_Sharpe_Weights")

    for r in dataframe_to_rows(
        max_sharpe.to_frame("Weight"), index=True, header=True
    ):
        ws_max.append(r)

    auto_adjust_columns(ws_max)
    format_weights(ws_max)

    # =====================================================
    # ================= MIN VOL ===========================
    # =====================================================

    ws_min = wb.create_sheet("Min_Vol_Weights")

    for r in dataframe_to_rows(
        min_vol.to_frame("Weight"), index=True, header=True
    ):
        ws_min.append(r)

    auto_adjust_columns(ws_min)
    format_weights(ws_min)

    add_seasonality_sheet(wb, weekly_seasonality)
    wb.save(output_path)